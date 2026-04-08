from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.forms import inlineformset_factory
from .models import Pedido, Cliente, Producto, DetallePedido
from .forms import PedidoForm, ClienteForm, ProductoForm, DetallePedidoForm
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.utils import timezone
import datetime

# API JWT (se mantiene para uso externo/Postman)
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework_simplejwt.tokens import RefreshToken
from .serializers import LoginSerializer, UserSerializer


# ── Login / Logout HTML ───────────────────────────────────────────────────────

def login_view(request):
    # Si ya está autenticado, va directo al listado
    if request.user.is_authenticated:
        return redirect('listar_pedidos')

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            return redirect('listar_pedidos')
        else:
            messages.error(request, 'Usuario o contraseña incorrectos.')

    return render(request, 'users/login.html')


def logout_view(request):
    logout(request)
    return redirect('login')


# ── CRUD Pedidos (protegido con @login_required) ──────────────────────────────

@login_required(login_url='login')
def listar_pedidos(request):
    pedidos = Pedido.objects.all()
    return render(request, 'pedidos/listar_pedidos.html', {'pedidos': pedidos})


@login_required(login_url='login')
def crear_pedido(request):
    DetallePedidoFormSet = inlineformset_factory(
        Pedido, DetallePedido,
        form=DetallePedidoForm,
        extra=0,
        can_delete=True
    )
    if request.method == 'POST':
        form    = PedidoForm(request.POST)
        formset = DetallePedidoFormSet(request.POST, prefix='form')
        if form.is_valid() and formset.is_valid():
            pedido           = form.save(commit=False)
            pedido.save()
            formset.instance = pedido
            formset.save()

            # ── Descontar stock si el pedido NO está cancelado ──
            productos_bajo_stock = []
            if pedido.estado != 'cancelado':
                for detalle in pedido.detalles.all():
                    producto          = detalle.producto
                    producto.stock   -= detalle.cantidad
                    producto.save()
                    # Verificar si quedó con stock <= 5
                    if producto.stock <= 5:
                        productos_bajo_stock.append(producto.nombre)

            # Mensaje de éxito
            messages.success(request, 'Pedido creado correctamente.')

            # Alerta de stock bajo
            if productos_bajo_stock:
                nombres = ', '.join(productos_bajo_stock)
                messages.warning(request, f'⚠ Stock bajo: {nombres}. Necesita reabastecimiento.')

            return redirect('listar_pedidos')
    else:
        form    = PedidoForm()
        formset = DetallePedidoFormSet(prefix='form')

    productos = Producto.objects.filter(stock__gt=5)
    return render(request, 'pedidos/crear_pedido.html', {
        'form':      form,
        'formset':   formset,
        'productos': productos,
    })


@login_required(login_url='login')
def ver_pedido(request, pk):
    pedido = get_object_or_404(Pedido, pk=pk)
    return render(request, 'pedidos/ver_pedido.html', {'pedido': pedido})


@login_required(login_url='login')
def actualizar_pedido(request, pk):
    pedido = get_object_or_404(Pedido, pk=pk)

    if pedido.esta_bloqueado():
        messages.error(request, 'Este pedido está cancelado y no puede editarse.')
        return redirect('listar_pedidos')

    puede_editar_completo = pedido.puede_editar_completo()

    DetallePedidoFormSet = inlineformset_factory(
        Pedido, DetallePedido,
        form=DetallePedidoForm,
        extra=0,
        can_delete=True
    )

    # ── Stock disponible real = stock actual + cantidad del pedido ──
    detalles_actuales = pedido.detalles.select_related('producto').all()
    stock_disponible_real = {
        d.producto.id: {
            'stock_actual':   d.producto.stock,
            'cant_original':  d.cantidad,
            'max_disponible': d.producto.stock + d.cantidad,
            'nombre':         d.producto.nombre,
            'precio':         float(d.producto.precio),
        }
        for d in detalles_actuales
    }

    # Productos para el select (stock > 5 + los del pedido actual)
    productos_qs = Producto.objects.filter(stock__gt=5)
    ids_pedido   = [d.producto.id for d in detalles_actuales]
    if ids_pedido:
        productos_qs = (productos_qs | Producto.objects.filter(pk__in=ids_pedido)).distinct()

    if request.method == 'POST':

        if not puede_editar_completo:
            nuevo_estado    = request.POST.get('estado')
            estados_validos = ['pendiente', 'procesando', 'enviado', 'entregado', 'cancelado']
            if nuevo_estado not in estados_validos:
                messages.error(request, 'Estado no válido.')
                return redirect('listar_pedidos')

            estado_anterior = pedido.estado
            pedido.estado   = nuevo_estado
            pedido.save()

            if nuevo_estado == 'cancelado' and estado_anterior != 'cancelado':
                for detalle in pedido.detalles.all():
                    producto        = detalle.producto
                    producto.stock += detalle.cantidad
                    producto.save()
                messages.success(request, 'Pedido cancelado. Stock restaurado.')
            else:
                messages.success(request, f'Estado actualizado a "{pedido.get_estado_display()}".')
            return redirect('listar_pedidos')

        estado_anterior = pedido.estado
        cantidades_anteriores = {d.producto.id: d.cantidad for d in detalles_actuales}

        form    = PedidoForm(request.POST, instance=pedido)
        formset = DetallePedidoFormSet(request.POST, instance=pedido, prefix='form')

        if form.is_valid() and formset.is_valid():
            nuevo_estado = form.cleaned_data['estado']

            errores_stock = []
            if nuevo_estado != 'cancelado':
                for f in formset:
                    if f.cleaned_data.get('DELETE', False):
                        continue
                    if not f.cleaned_data.get('producto'):
                        continue
                    producto         = f.cleaned_data['producto']
                    cant_nueva       = f.cleaned_data.get('cantidad', 0)
                    cant_anterior    = cantidades_anteriores.get(producto.id, 0)
                    max_disp         = producto.stock + cant_anterior
                    if cant_nueva > max_disp:
                        errores_stock.append(
                            f'"{producto.nombre}": máximo disponible es {max_disp} '
                            f'(stock: {producto.stock} + {cant_anterior} del pedido).'
                        )

            if errores_stock:
                for error in errores_stock:
                    messages.error(request, error)
                return render(request, 'pedidos/actualizar_pedido.html', {
                    'form':                   form,
                    'formset':                formset,
                    'pedido':                 pedido,
                    'productos':              productos_qs,
                    'stock_disponible_real':  stock_disponible_real,
                    'puede_editar_completo':  puede_editar_completo,
                })

            pedido_actualizado = form.save(commit=False)

            if nuevo_estado == 'cancelado' and estado_anterior != 'cancelado':
                for detalle in pedido.detalles.all():
                    producto        = detalle.producto
                    producto.stock += detalle.cantidad
                    producto.save()
                pedido_actualizado.save()
                formset.save()
                messages.success(request, 'Pedido cancelado. Stock restaurado.')
                return redirect('listar_pedidos')

            pedido_actualizado.save()
            formset.save()

            productos_bajo_stock = []
            if estado_anterior == 'cancelado' and nuevo_estado != 'cancelado':
                for detalle in pedido.detalles.all():
                    producto          = detalle.producto
                    producto.stock   -= detalle.cantidad
                    producto.save()
                    if producto.stock <= 5:
                        productos_bajo_stock.append(producto.nombre)
            elif estado_anterior != 'cancelado' and nuevo_estado != 'cancelado':
                for detalle in pedido.detalles.all():
                    producto      = detalle.producto
                    cant_anterior = cantidades_anteriores.get(producto.id, 0)
                    diferencia    = detalle.cantidad - cant_anterior
                    if diferencia != 0:
                        producto.stock -= diferencia
                        producto.save()
                    if producto.stock <= 5:
                        productos_bajo_stock.append(producto.nombre)

            messages.success(request, 'Pedido actualizado correctamente.')
            if productos_bajo_stock:
                nombres = ', '.join(productos_bajo_stock)
                messages.warning(request, f'⚠ Stock bajo: {nombres}. Necesita reabastecimiento.')
            return redirect('listar_pedidos')

    else:
        form    = PedidoForm(instance=pedido)
        formset = DetallePedidoFormSet(instance=pedido, prefix='form')

    return render(request, 'pedidos/actualizar_pedido.html', {
        'form':                  form,
        'formset':               formset,
        'pedido':                pedido,
        'productos':             productos_qs,
        'stock_disponible_real': stock_disponible_real,
        'puede_editar_completo': puede_editar_completo,
    })
@login_required(login_url='login')
def eliminar_pedido(request, pk):
    pedido = get_object_or_404(Pedido, pk=pk)
    if request.method == 'POST':
        pedido.delete()
        messages.success(request, f'Pedido #{pk} eliminado correctamente.')
    return redirect('listar_pedidos')  # ← siempre redirige, sin render


# ── PDF Pedidos ───────────────────────────────────────────────
@login_required(login_url='login')
def exportar_pedidos_pdf(request):
    pedidos  = Pedido.objects.all().prefetch_related('detalles__producto')
    template = get_template('pedidos/pedidos_pdf.html')
    html     = template.render({'pedidos': pedidos})
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="pedidos.pdf"'
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Error al generar el PDF', status=500)
    return response

# ── PDF Clientes ──────────────────────────────────────────────
@login_required(login_url='login')
def exportar_clientes_pdf(request):
    clientes = Cliente.objects.all()
    template = get_template('clientes/clientes_pdf.html')
    html     = template.render({'clientes': clientes})
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="clientes.pdf"'
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Error al generar el PDF', status=500)
    return response

# ── PDF Productos ─────────────────────────────────────────────
@login_required(login_url='login')
def exportar_productos_pdf(request):
    productos = Producto.objects.all()
    template  = get_template('productos/productos_pdf.html')
    html      = template.render({'productos': productos})
    response  = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="productos.pdf"'
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Error al generar el PDF', status=500)
    return response

# ── Excel Pedidos ─────────────────────────────────────────────
@login_required(login_url='login')
def exportar_pedidos_excel(request):
    pedidos = Pedido.objects.all().prefetch_related('detalles__producto')
    wb      = openpyxl.Workbook()
    ws      = wb.active
    ws.title = 'Pedidos'

    # Estilos
    header_font    = Font(bold=True, color='FFFFFF')
    header_fill    = PatternFill('solid', fgColor='1a1a2e')
    center         = Alignment(horizontal='center')

    # Encabezados
    headers = ['ID', 'Cliente', 'Fecha', 'Estado', 'Productos', 'Total']
    for col, header in enumerate(headers, 1):
        cell           = ws.cell(row=1, column=col, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center

    # Datos
    for row, pedido in enumerate(pedidos, 2):
        productos_str = ', '.join(
            f"{d.producto.nombre} x{d.cantidad}" for d in pedido.detalles.all()
        )
        ws.cell(row=row, column=1, value=pedido.id)
        ws.cell(row=row, column=2, value=str(pedido.cliente))
        ws.cell(row=row, column=3, value=str(pedido.fecha_pedido))
        ws.cell(row=row, column=4, value=pedido.get_estado_display())
        ws.cell(row=row, column=5, value=productos_str)
        ws.cell(row=row, column=6, value=float(pedido.total()))

    # Ancho de columnas
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 15

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="pedidos.xlsx"'
    wb.save(response)
    return response

# ── Excel Clientes ────────────────────────────────────────────
@login_required(login_url='login')
def exportar_clientes_excel(request):
    clientes = Cliente.objects.all()
    wb       = openpyxl.Workbook()
    ws       = wb.active
    ws.title = 'Clientes'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1a1a2e')
    center      = Alignment(horizontal='center')

    headers = ['ID', 'Nombre', 'Correo', 'Dirección', 'Teléfono']
    for col, header in enumerate(headers, 1):
        cell           = ws.cell(row=1, column=col, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center

    for row, cliente in enumerate(clientes, 2):
        ws.cell(row=row, column=1, value=cliente.id)
        ws.cell(row=row, column=2, value=cliente.nombre)
        ws.cell(row=row, column=3, value=cliente.correo)
        ws.cell(row=row, column=4, value=cliente.direccion)
        ws.cell(row=row, column=5, value=cliente.telefono)

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 15

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="clientes.xlsx"'
    wb.save(response)
    return response

# ── Excel Productos ───────────────────────────────────────────
@login_required(login_url='login')
def exportar_productos_excel(request):
    productos = Producto.objects.all()
    wb        = openpyxl.Workbook()
    ws        = wb.active
    ws.title  = 'Productos'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1a1a2e')
    center      = Alignment(horizontal='center')

    headers = ['ID', 'Nombre', 'Precio', 'Stock', 'Estado Stock']
    for col, header in enumerate(headers, 1):
        cell           = ws.cell(row=1, column=col, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center

    for row, producto in enumerate(productos, 2):
        estado_stock = 'Crítico' if producto.stock <= 5 else 'OK'
        ws.cell(row=row, column=1, value=producto.id)
        ws.cell(row=row, column=2, value=producto.nombre)
        ws.cell(row=row, column=3, value=float(producto.precio))
        ws.cell(row=row, column=4, value=producto.stock)
        ws.cell(row=row, column=5, value=estado_stock)

        # Color rojo si stock crítico
        if producto.stock <= 5:
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = PatternFill('solid', fgColor='FFE0E0')

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="productos.xlsx"'
    wb.save(response)
    return response
# ── CRUD Clientes ─────────────────────────────────────────────

@login_required(login_url='login')
def listar_clientes(request):
    clientes = Cliente.objects.all()
    return render(request, 'clientes/listar_clientes.html', {'clientes': clientes})


@login_required(login_url='login')
def crear_cliente(request):
    if request.method == 'POST':
        form = ClienteForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cliente creado correctamente.')
            return redirect('listar_clientes')
    else:
        form = ClienteForm()
    return render(request, 'clientes/crear_cliente.html', {'form': form})


@login_required(login_url='login')
def editar_cliente(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    if request.method == 'POST':
        form = ClienteForm(request.POST, instance=cliente)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cliente actualizado correctamente.')
            return redirect('listar_clientes')
    else:
        form = ClienteForm(instance=cliente)
    return render(request, 'clientes/editar_cliente.html', {'form': form, 'cliente': cliente})


@login_required(login_url='login')
def eliminar_cliente(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    if request.method == 'POST':
        cliente.delete()
        messages.success(request, f'Cliente "{cliente.nombre}" eliminado correctamente.')
    return redirect('listar_clientes')



# ── CRUD Productos ────────────────────────────────────────────

@login_required(login_url='login')
def listar_productos(request):
    productos = Producto.objects.all()
    return render(request, 'productos/listar_productos.html', {'productos': productos})


@login_required(login_url='login')
def crear_producto(request):
    if request.method == 'POST':
        form = ProductoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Producto creado correctamente.')
            return redirect('listar_productos')
    else:
        form = ProductoForm()
    return render(request, 'productos/crear_producto.html', {'form': form})


@login_required(login_url='login')
def editar_producto(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    if request.method == 'POST':
        form = ProductoForm(request.POST, instance=producto)
        if form.is_valid():
            form.save()
            messages.success(request, 'Producto actualizado correctamente.')
            return redirect('listar_productos')
    else:
        form = ProductoForm(instance=producto)
    return render(request, 'productos/editar_producto.html', {'form': form, 'producto': producto})


@login_required(login_url='login')
def eliminar_producto(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    if request.method == 'POST':
        producto.delete()
        messages.success(request, f'Producto "{producto.nombre}" eliminado correctamente.')
    return redirect('listar_productos')
# ── API JWT (para Postman o frontend externo) ─────────────────────────────────

class LoginView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = LoginSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        user = serializer.validated_data['user']
        refresh = RefreshToken.for_user(user)
        return Response({
            'user':    UserSerializer(user).data,
            'access':  str(refresh.access_token),
            'refresh': str(refresh),
        }, status=status.HTTP_200_OK)


class LogoutView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            token = RefreshToken(request.data.get('refresh'))
            token.blacklist()
            return Response({'detail': 'Sesión cerrada.'})
        except Exception:
            return Response({'detail': 'Token inválido.'}, status=status.HTTP_400_BAD_REQUEST)


class ProfileView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        return Response(UserSerializer(request.user).data)